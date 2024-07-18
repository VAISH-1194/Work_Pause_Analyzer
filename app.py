from flask import Flask, request, render_template, send_file, Response
import pandas as pd
import os
import zipfile
from split import split_tables
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from werkzeug.utils import secure_filename
from io import BytesIO
import re

app = Flask(__name__)

@app.route('/')
def upload_form():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return 'No file part'
    
    files = request.files.getlist('file')
    if not files or all(file.filename == '' for file in files):
        return 'No selected file'

    output_files = []

    for file in files:
        df = pd.read_excel(file)

        def drop_specific_rows(df):
            df = df.astype(str)
            mask = df.apply(lambda row: row.str.startswith(('Total')).any(), axis=1)
            df = df[~mask]
            return df

        df = drop_specific_rows(df)

        def move_rows_to_end(df):
            condition = df.iloc[:, 1].astype(str).str.startswith(("Department", "Emp Code"))

            rows_to_move = df[condition]

            remaining_rows = df[~condition]

            result_df = pd.concat([remaining_rows, rows_to_move], ignore_index=True)

            df.iloc[:] = result_df

        move_rows_to_end(df)

        file_path = "ad_tb.xlsx"
        df.to_excel(file_path, index=False)

        df = pd.read_excel(file_path, header=1)
        df = df.drop([col for col in df.columns if col.startswith('Unnamed')], axis=1)

        columns_to_keep = ["Att. Date", "InTime", "OutTime", "Shift", "S. InTime", "S. OutTime", "Punch Records"]
        df = df[columns_to_keep]

        df = pd.DataFrame(df)
        df['Records'] = df['Punch Records']


        def change_name(records):
            if pd.isna(records):
                return records
            entries = records.split(',')
            entries = [entry.replace('BD', 'ED') for entry in entries]
            entries = [entry.replace('Main Entrance', 'ED').replace('Exit', 'ED') for entry in entries]

            return ', '.join(entries)
        df['Records'] = df['Records'].apply(change_name)


        def update_in_out_times(row):
            records = row["Punch Records"]
            
            if pd.isna(records):
                return pd.Series({'InTime': ' ', 'OutTime': ' '})
            
            entries = records.split(',')
            in_time_matches = re.findall(r"\d{2}:\d{2}", entries[0]) if entries else []
            in_time = in_time_matches[0] if in_time_matches else ' '
            
            last_entry = entries[-2] if len(entries) > 1 else ' '
            out_time_matches = re.findall(r"\d{2}:\d{2}", last_entry) if last_entry else []
            out_time = out_time_matches[0] if out_time_matches else ' '
            
            if 'out' not in last_entry:
                out_time += ", records missing"
            
            return pd.Series({'InTime': in_time, 'OutTime': out_time})

        df_filtered = df[~df['Att. Date'].str.startswith(('Emp Code', 'Department'))]

        df_filtered.loc[:, ['InTime', 'OutTime']] = df_filtered.apply(update_in_out_times, axis=1)


        df.update(df_filtered[['InTime', 'OutTime']])

        df['InTime'] = df['InTime'].fillna(' ')
        df['OutTime'] = df['OutTime'].fillna(' ')


        def remove_1st_entries(record):
            if pd.isna(record):
                return record

            entries = record.split(', ')
            filtered_entries = [entry for entry in entries if "1st" not in entry]
            return ', '.join(filtered_entries)

        df['Records'] = df['Records'].apply(remove_1st_entries)


        def filter_punch_records(record):
            if pd.isna(record):
                return record

            entries = record.split(',')

            valid_entries = [entry for entry in entries if ('in' in entry or 'out' in entry)]

            return ','.join(valid_entries)

        df['Records'] = df['Records'].apply(filter_punch_records)

        df['Corrected Records'] = df['Records']

        def calculate_duration(row):
            in_time_str = row['InTime']
            out_time_str = row['OutTime'].replace(", records missing", "").strip()
            
            if in_time_str == ' ' or out_time_str == ' ':
                return ' '
            
            try:
                in_time = datetime.strptime(in_time_str, '%H:%M')
                out_time = datetime.strptime(out_time_str, '%H:%M')
                
                if out_time < in_time:
                    out_time += pd.DateOffset(days=1)
                
                duration = out_time - in_time
                total_minutes = duration.total_seconds() / 60
                hours, minutes = divmod(total_minutes, 60)
                return f"{int(hours)}h {int(minutes)}m"
            except Exception as e:
                return ' '

        df['Total Duration'] = df.apply(calculate_duration, axis=1)

        df['Punch Records'].replace('NaN', pd.NA, inplace=True)
        df['Records'].replace('NaN', pd.NA, inplace=True)

        df['Employee Status'] = df['Records'].apply(lambda x: 'Present' if pd.notna(x) and x != '' else 'Absent')


        def update_status_based_on_records(records, punch_records):
            if pd.isna(records) or records.strip() == '':
                return 'Absent'

            entries = records.split(', ')

            if len(entries) == 1:
                return 'Punch records missing'

            if 'out' in entries[0]:
                return 'Punch records missing'

            if len(entries) % 2 != 0:
                return 'Punch records missing'

            for i in range(1, len(entries)):
                if ('in' in entries[i] and 'in' in entries[i-1]) or \
                ('out' in entries[i] and 'out' in entries[i-1]):
                    return 'Punch records missing'

            return 'Valid Records'

        def mark_columns_empty(row):
            words_to_check = ['Att. Date', 'InTime', 'OutTime', 'Shift', 'S. InTime', 'S. OutTime',
                            'Punch Records', 'Records']

            if row.name > 0:
                for word in words_to_check:
                    if pd.notna(row[word]) and any(word in str(cell) for cell in row):
                        row['Employee Status'] = " "
                        row['Records Status'] = " "
                        row['Break Time'] = " "
                        return row
            return row

        for col in ['Employee Status', 'Records Status', 'Break Time']:
            if col not in df.columns:
                df[col] = ""

        df = df.apply(mark_columns_empty, axis=1)

        df['Records Status'] = df.apply(lambda row: update_status_based_on_records(row['Records'], row['Punch Records']), axis=1)

        cols_to_check = ['Att. Date', 'InTime', 'OutTime', 'Shift', 'S. InTime', 'S. OutTime',
                        'Punch Records', 'Records']

        def check_nan_and_update_status(row, cols_to_check):
            if all(pd.isna(row[col]) for col in cols_to_check):
                row['Employee Status'] = " "
                row['Records Status'] = " "
                row['Break Time'] = " "
            return row

        df = df.apply(lambda row: check_nan_and_update_status(row, cols_to_check), axis=1)

        columns_order = [col for col in df.columns if col not in ['Employee Status', 'Records Status', 'Break Time']]
        columns_order += ['Employee Status', 'Records Status', 'Break Time']
        df = df[columns_order]

        df['Records_Dup'] = df['Records']

        def check_and_adjust_entries(records_dup):
            if pd.isna(records_dup):
                return '', 'N/A'

            entries = str(records_dup).split(', ')
            if len(entries) % 2 != 0:
                if entries[0].endswith("in(ED)") and entries[-1].endswith("in(ED)"):
                    entries.append('--:--:out(ED)')
                    return ', '.join(entries), 'Partially valid'

            for i in range(1, len(entries)):
                if (entries[i].endswith("in(ED)") and entries[i-1].endswith("in(ED)")) or (entries[i].endswith("out(ED)") and entries[i-1].endswith("out(ED)")):
                    return ', '.join(entries), 'Invalid Records'

            return ', '.join(entries), 'Present'

        df['Records_Dup'], df['Validity'] = zip(*df['Records_Dup'].apply(check_and_adjust_entries))

        df['Corrected Records'] = df['Records']

        def update_approx_break_time(row):
            if row['Employee Status'] == 'Absent':
                return 'N/A'
            elif row['Records Status'] == 'Valid Records':
                return ''
            elif row['Validity'] == 'Partially valid':
                return 'Partially valid'
            elif row['Validity'] == 'Invalid Records':
                return 'Invalid Records'
            else:
                return 'N/A'

        df['Approx. Break Time'] = df.apply(update_approx_break_time, axis=1)

        def remove_first_in_last_out(records):
            entries = records.split(', ')
            if len(entries) > 0 and entries[0].endswith('in(ED)'):
                entries.pop(0)
            if len(entries) > 0 and entries[-1].endswith('out(ED)'):
                entries.pop(-1)
            return ', '.join(entries)

        df['Records_Dup'] = df['Records_Dup'].apply(lambda x: remove_first_in_last_out(x) if pd.notna(x) else x)

        def calculate_break_time(record):
            entries = record.split(', ')
            total_break_time = 0

            if len(entries) % 2 != 0:
                return 'Invalid entry length'

            for i in range(1, len(entries), 2):
                out_time_str = entries[i - 1].split(':out(ED)')[0].strip()
                in_time_str = entries[i].split(':in(ED)')[0].strip()

                out_time_match = re.search(r'\d{2}:\d{2}', out_time_str)
                in_time_match = re.search(r'\d{2}:\d{2}', in_time_str)

                if out_time_match and in_time_match:
                    out_time = pd.to_datetime(out_time_match.group(), format='%H:%M')
                    in_time = pd.to_datetime(in_time_match.group(), format='%H:%M')
                    if in_time < out_time:
                        in_time += pd.Timedelta(days=1)
                    break_duration = in_time - out_time
                    total_break_time += break_duration.total_seconds() / 60
                else:
                    return 'Invalid time format'

            return int(total_break_time)

        def format_break_time(minutes):
            if isinstance(minutes, str):
                return minutes
            hours = minutes // 60
            mins = minutes % 60
            if hours > 0:
                return f"{hours} hr {mins} mins" if mins > 0 else f"{hours} hr"
            else:
                return f"{mins} mins"

        def final_update_approx_break_time(row):
            if 'Partially valid' in row['Approx. Break Time']:
                break_time_minutes = calculate_break_time(row['Records_Dup'])
                formatted_break_time = format_break_time(break_time_minutes)

                if row['Employee Status'] == 'Absent':
                    return 'N/A'
                elif row['Records Status'] == 'Valid Records':
                    return formatted_break_time
                elif 'Partially valid' in row['Approx. Break Time']:
                    return f"Partially valid, {formatted_break_time}"
                elif 'Invalid Records' in row['Approx. Break Time']:
                    return 'Invalid Records'
                else:
                    return formatted_break_time
            else:
                return row['Approx. Break Time']

        df['Approx. Break Time'] = df.apply(final_update_approx_break_time, axis=1)  

        def handle_invalid_entries(approx_break_time):
            if pd.isna(approx_break_time):
                return approx_break_time
            if "Partially valid, Invalid entry length" in approx_break_time:
                return "Invalid Entries, Punch missed"
            if re.search(r'-\d+', approx_break_time):
                return "Invalid Entries, Punch missed"
            return approx_break_time

        df['Approx. Break Time'] = df['Approx. Break Time'].apply(handle_invalid_entries)

        def remove_first_in_last_out(records):
            entries = records.split(', ')
            if len(entries) > 0 and entries[0].endswith('in(ED)'):
                entries.pop(0)
            if len(entries) > 0 and entries[-1].endswith('out(ED)'):
                entries.pop(-1)
            return ', '.join(entries)

        df['Records'] = df['Records'].apply(lambda x: remove_first_in_last_out(x) if pd.notna(x) else x)

        def calculate_break_time(row):
            if row['Employee Status'] == 'Absent':
                return 'N/A'

            if row['Employee Status'] == 'Present':
                entries = row['Records'].split(',')
                total_break_time = 0
                for i in range(1, len(entries), 2):
                    in_time_str = entries[i - 1].split()[-1]
                    out_time_str = entries[i].split()[-1]

                    in_time_match = re.search(r'\d{2}:\d{2}', in_time_str)
                    out_time_match = re.search(r'\d{2}:\d{2}', out_time_str)

                    if in_time_match and out_time_match:
                        in_time = pd.to_datetime(in_time_match.group(), format='%H:%M')
                        out_time = pd.to_datetime(out_time_match.group(), format='%H:%M')
                        break_duration = out_time - in_time
                        total_break_time += break_duration.total_seconds() / 60

                    if row['Records Status'] == 'Punch records missing':
                        return 'N/A'

                return int(total_break_time)
            return 0

        def format_break_time(minutes):
            if minutes == 'N/A':
                return minutes
            hours = minutes // 60
            mins = minutes % 60
            if hours > 0:
                return f"{hours} hr {mins} mins" if mins > 0 else f"{hours} hr"
            else:
                return f"{mins} mins"

        df['Break Time'] = df.apply(calculate_break_time, axis=1)
        df['Break Time'] = df['Break Time'].apply(format_break_time)

        def update_break_time_for_missing_records(row):
            if row['Records Status'] == 'Punch records missing':
                return 'N/A'
            return row['Break Time']

        df['Break Time'] = df.apply(update_break_time_for_missing_records, axis=1)

        df['Punch Records'].replace('NaN', pd.NA, inplace=True)
        df['Records'].replace('NaN', pd.NA, inplace=True)

        df['Employee Status'] = df['Records'].apply(lambda x: 'Present' if pd.notna(x) and x != '' else 'Absent')

        columns_to_drop = ['Records_Dup','Validity','Records']

        df.columns = df.columns.str.strip()

        df.drop(columns=columns_to_drop, errors='ignore', inplace=True)

        def mark_columns_empty(row):
            words_to_check = ['Att. Date', 'InTime', 'OutTime', 'Shift', 'S. InTime', 'S. OutTime',
                            'Punch Records', 'Corrected Records']

            if row.name > 0:
                for word in words_to_check:
                    if word in row and pd.notna(row[word]) and any(word in str(cell) for cell in row):
                        row['Employee Status'] = " "
                        row['Records Status'] = " "
                        row['Break Time'] = " "
                        return row
            return row

        def update_status_based_on_records(records, punch_records):
            if pd.isna(records) or records.strip() == '':
                if pd.isna(punch_records) or punch_records.strip() == '':
                    return 'Absent'
                else:
                    return 'Punch records missing'
            entries = records.split(', ')

            if len(entries) == 1:
                return 'Punch records missing'

            if 'out' in entries[0]:
                return 'Punch records missing'

            if len(entries) % 2 != 0:
                return 'Punch records missing'

            for i in range(1, len(entries)):
                if ('in' in entries[i] and 'in' in entries[i-1]) or \
                ('out' in entries[i] and 'out' in entries[i-1]):
                    return 'Punch records missing'

            return 'Valid Records'

        if 'Records' in df.columns and 'Punch Records' in df.columns:
            df['Records Status'] = df.apply(lambda row: update_status_based_on_records(row['Records'], row['Punch Records']), axis=1)

        cols_to_check = ['Att. Date', 'InTime', 'OutTime', 'Shift', 'S. InTime', 'S. OutTime',
                        'Punch Records', 'Records']

        def check_nan_and_update_status(row, cols_to_check):
            if any(str(row[col]).strip().lower().startswith(('department', 'emp code')) for col in row.index if pd.notna(row[col])):
                row['Employee Status'] = " "
                row['Records Status'] = " "
                row['Break Time'] = " "
                row['Approx. Break Time'] = " "
                row['Total Duration'] = " "
            else:
                if 'Punch Records' in row and pd.notna(row['Punch Records']) and row['Punch Records'].strip() != '':
                    row['Employee Status'] = "Present"
                else:
                    row['Employee Status'] = "Absent"
            return row

        df = df.apply(lambda row: check_nan_and_update_status(row, cols_to_check), axis=1)

        df = df.apply(mark_columns_empty, axis=1)
        columns_order = [col for col in df.columns if col not in ['Corrected Records', 'Records Status','Total Duration', 'Employee Status', 'Break Time', 'Approx. Break Time']]
        columns_order += ['Corrected Records', 'Records Status', 'Total Duration', 'Employee Status', 'Break Time', 'Approx. Break Time']
        df = df[columns_order]

        def should_drop_row(row):
            first_cell_value = str(row.iloc[0])
            return first_cell_value.startswith(('Total'))

        df = df[~df.apply(should_drop_row, axis=1)]

        leave_dates = []
        total_leaves = 0

        for idx, row in df.iterrows():
            att_date = pd.to_datetime(row['Att. Date'], errors='coerce')
            day_of_week = att_date.weekday() if pd.notnull(att_date) else None

            if row['Employee Status'] == 'Absent' and day_of_week is not None and day_of_week < 5:
                leave_dates.append(row['Att. Date'])
                total_leaves += 1

        leave_dates_row = {
            'Att. Date': 'Leave Dates:',
            'InTime': ', '.join(leave_dates) if leave_dates else '',
            'OutTime': 'No. of leaves:',
            'Shift': total_leaves,
            'S. InTime': '',
            'S. OutTime': '',
            'Punch Records': '',
            'Corrected Records': '',
            'Records Status': '',
            'Total Duration': '',
            'Employee Status': '',
            'Break Time': '',
            'Approx. Break Time': ''
        }

        leave_dates_df = pd.DataFrame([leave_dates_row])

        try:
            emp_code_index = df.index[df['Att. Date'].astype(str).str.contains('Emp Code:', na=False)].tolist()[0]
            df = pd.concat([df.iloc[:emp_code_index + 1], leave_dates_df, df.iloc[emp_code_index + 1:]], ignore_index=True)
        except IndexError:
            print("Error: 'Emp Code:' not found in the 'Att. Date' column")


        temp_file_path = 'temp_file.xlsx'
        df.to_excel(temp_file_path, index=False)
        wb = load_workbook(temp_file_path)
        ws = wb.active

        aqua_fill = PatternFill(start_color="CCDCF8", end_color="CCDCF8", fill_type="solid")
        f8c9eb_fill = PatternFill(start_color="F9D3EE", end_color="F9D3EE", fill_type="solid")
        head_fill = PatternFill(start_color="DBFBEA", end_color="DBFBEA", fill_type="solid")
        specific_fill = PatternFill(start_color="D9C5E9", end_color="D9C5E9", fill_type="solid")
        baabcd_fill = PatternFill(start_color="BAABCD", end_color="BAABCD", fill_type="solid")

        thin_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))

        att_date_col_idx = df.columns.get_loc('Att. Date') + 1
        break_time_col_idx = df.columns.get_loc('Break Time') + 1
        approx_break_time_col_idx = df.columns.get_loc('Approx. Break Time') + 1

        header_keywords = ["Employee Name :", "Department:", "Emp Code:", "Leave Dates:", "No. of leaves:"]

        def fill_specific_cells(ws, keywords, header_fill, value_fill):
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    if cell.value in keywords:
                        cell.fill = header_fill
                        cell.border = thin_border
                        if cell.value == "Employee Name :":
                            related_cell = cell.offset(column=3) 
                        else:
                            related_cell = cell.offset(column=1)
                        related_cell.fill = value_fill
                        related_cell.border = thin_border

        for cell in ws[1]:
            cell.fill = head_fill

        fill_specific_cells(ws, header_keywords, specific_fill, baabcd_fill)

        for row in ws.iter_rows(min_row=2):
            if row[att_date_col_idx - 1].value and row[0].value not in header_keywords:
                row[break_time_col_idx - 1].fill = aqua_fill
                row[break_time_col_idx - 1].border = thin_border
                row[approx_break_time_col_idx - 1].fill = f8c9eb_fill
                row[approx_break_time_col_idx - 1].border = thin_border

        header_row_height = 33.60  
        data_row_height = 33.60 

        ws.row_dimensions[1].height = header_row_height

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = data_row_height

        column_widths = {
            'InTime': 14.29,
            'OutTime': 22.56,
            'Shift': 14.29,
            'S. InTime': 14.29,
            'S. OutTime': 14.29,
            'Punch Records': 38.57,
            'Corrected Records': 38.57,
            'Approx. Break Time': 38.57,
            'Break Time': 16.00,
            'Total Duration': 16.00,
            'Att. Date': 21.44,
            'Employee Status': 21.44,
            'Records Status': 21.44
        }

        for col_name, width in column_widths.items():
            if col_name in df.columns:
                col_idx = df.columns.get_loc(col_name) + 1
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                ws.column_dimensions[col_letter].width = width

        alignment = Alignment(horizontal='center', vertical='center')
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = alignment

    #     employee_name = None
    #     for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    #         for cell in row:
    #             if cell.value == "Employee Name :":
    #                 employee_name = cell.offset(column=3).value
    #                 break
    #         if employee_name:
    #             break

    #     if not employee_name:
    #         employee_name = "Unnamed_Employee"

    #     output_file_name = f"{employee_name}.xlsx"
    #     wb.save(output_file_name)
    #     output_files.append(output_file_name)

    # zip_filename = 'processed_files.zip'
    # with zipfile.ZipFile(zip_filename, 'w') as zipf:
    #     for output_file in output_files:
    #         if os.path.exists(output_file):
    #             zipf.write(output_file)
    #             os.remove(output_file)  

    # return send_file(zip_filename, download_name=zip_filename, as_attachment=True)



     #Zip_file change response
     
        employee_name = None
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value == "Employee Name :":
                    employee_name = cell.offset(column=3).value
                    break
            if employee_name:
                break

        if not employee_name:
            employee_name = "Unnamed_Employee"

        output_file_name = f"{employee_name}.xlsx"
        wb.save(output_file_name)
        output_files.append(output_file_name)

    zip_filename = 'processed_files.zip'
    with zipfile.ZipFile(zip_filename, 'w') as zipf:
        for output_file in output_files:
            if os.path.exists(output_file):
                zipf.write(output_file)
                os.remove(output_file)

    with open(zip_filename, 'rb') as f:
        zip_data = f.read()

    os.remove(zip_filename)

    response = Response(
        zip_data,
        mimetype='application/zip',
        headers={'Content-Disposition': 'attachment; filename=processed_files.zip'}
    )

    return response




@app.route('/split', methods=['POST'])
def split_file():
    if 'file' not in request.files:
        return 'No file part'
    
    file = request.files['file']
    if file.filename == '':
        return 'No selected file'
    
    filename = secure_filename(file.filename)
    file.save(filename)
    df = pd.read_excel(file, header=None)

    tables = split_tables(df)

    output_dir = 'split_tables'
    os.makedirs(output_dir, exist_ok=True)
    output_files = []

    for i, (headers, table) in enumerate(tables):
        file_name = f'table_{i+1}.xlsx'
        file_path = os.path.join(output_dir, file_name)
        output_files.append(file_path)

        with pd.ExcelWriter(file_path) as writer:
            for j, header in enumerate(headers):
                header_df = pd.DataFrame([header])
                header_df.to_excel(writer, index=False, header=False, startrow=j)
            table.to_excel(writer, index=False, header=False, startrow=len(headers))

    # zip_filename = 'split_files.zip'
    # with zipfile.ZipFile(zip_filename, 'w') as zipf:
    #     for output_file in output_files:
    #         zipf.write(output_file, os.path.basename(output_file))
    
    # return send_file(zip_filename, as_attachment=True)



    #Zip_file change response

        zip_filename = 'split_files.zip'
    with zipfile.ZipFile(zip_filename, 'w') as zipf:
        for output_file in output_files:
            zipf.write(output_file, os.path.basename(output_file))

    with open(zip_filename, 'rb') as f:
        zip_data = f.read()

    os.remove(zip_filename)
    for output_file in output_files:
        os.remove(output_file)

    response = Response(
        zip_data,
        mimetype='application/zip',
        headers={'Content-Disposition': 'attachment; filename=split_files.zip'}
    )

    return response

if __name__ == "__main__":
    app.run(debug=True)
    